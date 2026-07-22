import openpyxl
import os

file_path = r"C:\Users\JDURAN1\Desktop\Proyecto Sacos Vacíos\SALDOS SACOS VACIOS.xlsx"
print(f"File exists: {os.path.exists(file_path)}")

if os.path.exists(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("Sheets:", wb.sheetnames)
        sheet = wb.active
        print("Active sheet title:", sheet.title)
        print("First 15 rows:")
        for r in range(1, 16):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")
    except Exception as e:
        print("Error reading excel:", e)
